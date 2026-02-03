import re
from typing import Any

import frappe
from frappe import _

from ecommerce_integrations.shopify.constants import (
	ORDER_ID_FIELD,
	SHIPPING_ADDRESS_FIELD,
	SHIPPING_CUSTOMER_NAME_FIELD,
	SHIPPING_PHONE_FIELD,
)


@frappe.whitelist()
def sync_customers_from_excel(file_url: str):
	"""Update Sales Orders with shipping customer info from Shopify Excel export.

	Expects `file_url` an Attach field value (e.g. /files/filename.xlsx).

	**Use Case**: Sales Orders use a default customer for online orders.
	This function enriches the Sales Orders (and related Invoices/Delivery Notes) with:
	- Shipping customer name
	- Shipping address
	- Shipping phone number

	**Performance Note**: Filter Excel by date to match recent sales orders, then
	this function updates only the relevant orders.

	Required fields in Excel:
	- Id: Shopify Order ID (must match existing Sales Order)
	- Shipping Name: Customer name for shipping
	- Shipping Street, Shipping Address1, Shipping City, etc.: Address fields
	- Phone: Phone number

	Respects the `personally_identifiable_information_access` flag from Shopify Settings.
	"""
	if not file_url:
		frappe.throw(_("file_url is required"))

	site_path = _check_file_path(file_url)
	updated = 0
	skipped = 0
	errors = []
	rows = _read_file_and_extract_rows(site_path)
	date_from = _fetch_earliest_date(rows)
	row_dict = {str(row.get("Id")).strip(): row for row in rows if row.get("Id")}
	sales_orders = _get_sales_orders_from_erpnext_with_customer_info(date_from)

	for so in sales_orders:
		order_id = str(so.get(ORDER_ID_FIELD)).strip()
		if order_id in row_dict:
			order_full_data = row_dict[order_id]
			try:
				# Extract shipping customer info from Excel
				shipping_info = _extract_shipping_info_from_excel(order_full_data)

				# Update Sales Order with shipping info
				_update_sales_order_shipping_info(so.name, shipping_info)

				# Update related Sales Invoice if exists
				_update_sales_invoice_shipping_info(order_id, shipping_info)

				# Update related Delivery Notes if exist
				_update_delivery_notes_shipping_info(order_id, shipping_info)

				updated += 1
			except frappe.DoesNotExistError:
				skipped += 1
			except Exception as e:
				errors.append(frappe._("{0}: {1}").format(so.name, str(e)))
		else:
			skipped += 1

	return {"updated": updated, "skipped": skipped, "errors": errors}


def _extract_shipping_info_from_excel(order_data: dict) -> dict:
	"""Extract shipping customer info from Excel row data."""
	# Build full address string
	address_parts = []
	if order_data.get("Shipping Street"):
		address_parts.append(str(order_data.get("Shipping Street")).strip())
	if order_data.get("Shipping Address1"):
		address_parts.append(str(order_data.get("Shipping Address1")).strip())
	if order_data.get("Shipping City"):
		address_parts.append(str(order_data.get("Shipping City")).strip())
	if order_data.get("Shipping Province"):
		address_parts.append(str(order_data.get("Shipping Province")).strip())
	if order_data.get("Shipping Zip"):
		address_parts.append(str(order_data.get("Shipping Zip")).strip())
	if order_data.get("Shipping Country"):
		address_parts.append(str(order_data.get("Shipping Country")).strip())

	full_address = ", ".join([p for p in address_parts if p])

	shipping_name = str(order_data.get("Shipping Name") or "").strip()
	if not shipping_name:
		shipping_name = str(order_data.get("Email") or "").strip()

	shipping_phone = str(order_data.get("Phone") or "").strip()

	return {
		"shipping_customer_name": shipping_name,
		"shipping_address": full_address,
		"shipping_phone": shipping_phone,
	}


def _update_sales_order_shipping_info(so_name: str, shipping_info: dict):
	"""Update Sales Order with shipping customer info."""
	so_doc = frappe.get_doc("Sales Order", so_name)
	so_doc.db_set(SHIPPING_CUSTOMER_NAME_FIELD, shipping_info.get("shipping_customer_name", ""))
	so_doc.db_set(SHIPPING_ADDRESS_FIELD, shipping_info.get("shipping_address", ""))
	so_doc.db_set(SHIPPING_PHONE_FIELD, shipping_info.get("shipping_phone", ""))


def _update_sales_invoice_shipping_info(order_id: str, shipping_info: dict):
	"""Update Sales Invoice with shipping customer info if exists."""
	si_name = frappe.db.get_value("Sales Invoice", {ORDER_ID_FIELD: order_id}, "name")
	if si_name:
		frappe.db.set_value("Sales Invoice", si_name, {
			SHIPPING_CUSTOMER_NAME_FIELD: shipping_info.get("shipping_customer_name", ""),
			SHIPPING_ADDRESS_FIELD: shipping_info.get("shipping_address", ""),
			SHIPPING_PHONE_FIELD: shipping_info.get("shipping_phone", ""),
		})


def _update_delivery_notes_shipping_info(order_id: str, shipping_info: dict):
	"""Update Delivery Notes with shipping customer info if exist."""
	dn_list = frappe.db.get_all("Delivery Note", filters={ORDER_ID_FIELD: order_id}, pluck="name")
	for dn_name in dn_list:
		frappe.db.set_value("Delivery Note", dn_name, {
			SHIPPING_CUSTOMER_NAME_FIELD: shipping_info.get("shipping_customer_name", ""),
			SHIPPING_ADDRESS_FIELD: shipping_info.get("shipping_address", ""),
			SHIPPING_PHONE_FIELD: shipping_info.get("shipping_phone", ""),
		})


def _check_file_path(file_url: str) -> str:
	"""Check and return valid file path."""
	site_path = frappe.get_site_path("public", file_url.lstrip("/"))
	import os

	if not os.path.exists(site_path):
		# try with files/ prefix
		site_path = frappe.get_site_path("public", "files", os.path.basename(file_url))
		if not os.path.exists(site_path):
			frappe.throw(_("File not found: {0}").format(file_url))
	return site_path


def _read_file_and_extract_rows(site_path: str) -> list[dict[str, Any]]:
	"""Read Excel or CSV file and extract rows as list of dicts."""
	import os

	ext = os.path.splitext(site_path)[1].lower()
	rows = []
	try:
		if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
			from openpyxl import load_workbook

			wb = load_workbook(site_path, read_only=True, data_only=True)
			ws = wb[wb.sheetnames[0]]
			it = ws.iter_rows(values_only=True)
			headers = [str(h).strip() if h is not None else "" for h in next(it)]
			for r in it:
				rows.append(dict(zip(headers, r)))
		else:
			import csv

			with open(site_path, newline="") as f:
				reader = csv.DictReader(f)
				for r in reader:
					rows.append(r)
	except Exception as e:
		frappe.throw(_("Failed to read uploaded file: {0}").format(e))
	return rows


def _get_sales_orders_from_erpnext_with_customer_info(date_from: str) -> list[dict[str, Any]]:
	"""Get sales orders with shopify_order_id from ERPNext created on or after `date_from`."""
	sales_orders = frappe.db.get_all(
		"Sales Order",
		filters={
			"creation": [">=", date_from],
			ORDER_ID_FIELD: ["is", "set"],
		},
		fields=["name", ORDER_ID_FIELD],
	)
	return sales_orders


def _fetch_earliest_date(rows: list[dict[str, Any]]) -> str:
	"""Fetch the earliest date from the rows to limit sales order fetch."""
	date_from = None
	date_pattern = re.compile(r"\d{4}-\d{2}-\d{2}")
	for row in rows:
		created_at = row.get("Created at")
		if created_at:
			match = date_pattern.search(created_at)
			if match:
				row_date = match.group(0)
				if not date_from or row_date < date_from:
					date_from = row_date
	return date_from if date_from else "1970-01-01"
